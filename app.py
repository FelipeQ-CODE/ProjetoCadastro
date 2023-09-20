import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl,xlrd
import pathlib
from openpyxl import Workbook


#Setando a aparencia padrao do sistema

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


#Classe principal do sistema ele chama as funçoes para iniciar o programa
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.aparencia()
        self.todo_sistema()
        
    #layout do APP     tamanho e Nome do projeto
    def layout_config(self):
        self.title("Sistema de Gestão de Cadastro")
        self.geometry("800x600")
        
    #Aparencia
    def aparencia(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema",bg_color="transparent",text_color=['#000',"#fff"]).place(x=50,y=500)
        
        #Tema de cores Light ,Dark;
        self.opt_apm = ctk.CTkOptionMenu(self,values=["Light","Dark","System"], command=self.change_apm).place(x=50, y=530)
    
    #Faixa azul do app    
    def todo_sistema(self):
        frame =ctk.CTkFrame(self, width=800,height=50,corner_radius=0,bg_color="teal",fg_color="teal").place(x=0,y=10)    
    #titulo
        title=ctk.CTkLabel(frame,text="Sistema de Gestao de Clientes", font=("Century Gothic bold",24),text_color="#fff").place(x=190,y=10)
    #Subtitulo
        span= ctk.CTkLabel(self,text="Por favor ,preencha o formulario !", font=("Century Gothic bold",16),text_color=["#000","#fff"]).place(x=50 , y=70)
         
         
        ficheiro =pathlib.Path("Clientes.xlsx")
       
        if ficheiro.exists():
             pass
        else:
                ficheiro=Workbook()
                folha= ficheiro.active
                folha['A1']="Nome Completo"
                folha['B1']="Contato"
                folha['C1']="Idade"
                folha['D1']="Genero"
                folha['E1']="Endereços"
                folha['F1']="Observaçoes"
                    
                ficheiro.save("Clientes.xlsx")
                
    #definindo a função o botão
        import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl,xlrd
import pathlib
from openpyxl import Workbook


#Setando a aparencia padrao do sistema

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


#Classe principal do sistema ele chama as funçoes para iniciar o programa
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.aparencia()
        self.todo_sistema()
        
    #layout do APP     tamanho e Nome do projeto
    def layout_config(self):
        self.title("Sistema de Gestão de Cadastro")
        self.geometry("800x600")
        
    #Aparencia
    def aparencia(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema",bg_color="transparent",text_color=['#000',"#fff"]).place(x=50,y=500)
        
        #Tema de cores Light ,Dark;
        self.opt_apm = ctk.CTkOptionMenu(self,values=["Light","Dark"], command=self.change_apm).place(x=50, y=530)
    
    #Faixa azul do app    
    def todo_sistema(self):
        frame =ctk.CTkFrame(self, width=800,height=50,corner_radius=0,bg_color="Cyan",fg_color="Cyan")
        frame.place(x=0,y=10)    
    #titulo
        title=ctk.CTkLabel(frame,text="Sistema de Gestao de Clientes", font=("Century Gothic bold",24),text_color="#000").place(x=190,y=10)
    #Subtitulo
        span= ctk.CTkLabel(self,text="Por favor ,preencha o formulario !", font=("Century Gothic bold",16),text_color=["#000","#fff"]).place(x=50 , y=70)
         
         
        ficheiro =pathlib.Path("Clientes.xlsx")
       
        if ficheiro.exists():
             pass
        else:
                ficheiro=Workbook()
                folha= ficheiro.active
                folha['A1']="Nome Completo"
                folha['B1']="Contato"
                folha['C1']="Idade"
                folha['D1']="Genero"
                folha['E1']="Endereços"
                folha['F1']="Observaçoes"
                    
                ficheiro.save("Clientes.xlsx")
                
    #definindo a função o botão
        def submit():
            #pegar os dados dos entrys
            nome = nome_value.get()
            contact= contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            adrress= address_entry.get()
            obs = obs_entry.get(0.0,END)
            
        #Se o usuario nao preencher todos os dados
            if (nome =="" or contact =="" or age=="" or adrress==""):
                messagebox.showerror("Sistema","Erro!\n Por favor preencha todos os dados!")
            else:
        #Recebendo dados no nosso Excel ou banco de dados!
                ficheiro = openpyxl.load_workbook('Clientes.xlsx')
                
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row + 1, value=nome)
                folha.cell(column=2, row=folha.max_row, value=contact)
                folha.cell(column=3, row=folha.max_row, value=age)
                folha.cell(column=4, row=folha.max_row, value=gender)
                folha.cell(column=5, row=folha.max_row, value=adrress)
                folha.cell(column=6, row=folha.max_row, value=obs)
                
                ficheiro.save(r"Clientes.xlsx")
                messagebox.showinfo("Sistema","Dados Cadastrados com sucesso!")
            
            
        def clear():
            nome_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            obs_entry.delete(0.0,END)
        
    #Variaveis de Texto
        nome_value=StringVar()
        contact_value=StringVar()
        age_value=StringVar()
        address_value=StringVar()
        
        
    #Entry Caixa de dialogo para informação. obs caixa em branco
        nome_entry= ctk.CTkEntry(self,width=350,textvariable=nome_value,font=("Century Gohtic bold",16),fg_color="transparent")    
        contact_entry= ctk.CTkEntry(self,width=200,textvariable=contact_value,font=("Century Gohtic bold",16),fg_color="transparent")
        age_entry= ctk.CTkEntry(self,width=150,textvariable=age_value,font=("Century Gohtic bold",16),fg_color="transparent")
        address_entry= ctk.CTkEntry(self,width=200,textvariable=address_value,font=("Century Gohtic bold",16),fg_color="transparent")        
    
    #Combobox caixa de genero com  3 opçoes
        gender_combobox = ctk.CTkComboBox(self,values=["Masculino","Feminino","Trasgenero"],font=("Centurty Gothic bold",14),width=150)
        gender_combobox.set("Masculino")
        
    #Entrada de Obcervaçoes
        obs_entry=ctk.CTkTextbox(self,width=450,height=150,font=("Arial",18),border_color="#aaa",border_width=2 ,fg_color="transparent")
        
    #LABEL Caixa de texto titulos
        lb_nome= ctk.CTkLabel(self,text="Nome completo:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_contact= ctk.CTkLabel(self,text="Contato:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_age= ctk.CTkLabel(self,text="Idade:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_gander= ctk.CTkLabel(self,text="Género:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_address= ctk.CTkLabel(self,text="Endereço:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_obs= ctk.CTkLabel(self,text="Observações:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        
     #BOTOES , tamanho cor com interação 
        btn_submit= ctk.CTkButton(self,text="Salvar dados".upper(),command=submit,fg_color="#151",hover_color="#131").place(x=300,y=420)
        btn_submit= ctk.CTkButton(self,text="Limpar campos".upper(),command=clear,fg_color="#555",hover_color="#333").place(x=500,y=420)
        
        
    #Posicionando os elementos na tela
        lb_nome.place(x=50 ,y=120)
        nome_entry.place(x=50 ,y=150)
        
        lb_contact.place(x=450,y=120)
        contact_entry.place(x=450,y=150)
        
        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)
        
        lb_gander.place(x=500,y=190)
        gender_combobox.place(x=500,y=220)
        
        lb_address.place(x=50,y=190)
        address_entry.place(x=50,y=220)
        
        lb_obs.place(x=50,y=260)
        obs_entry.place(x=180 ,y=260)
        
        
    def change_apm(self,nova_aparencia_mode):
        ctk.set_appearance_mode(nova_aparencia_mode)

            
        def clear():
            nome_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            obs_entry.delete(0.0,END)
        
    #Variaveis de Texto
        nome_value=StringVar()
        contact_value=StringVar()
        age_value=StringVar()
        address_value=StringVar()
        
        
    #Entry Caixa de dialogo para informação. obs caixa em branco
        nome_entry= ctk.CTkEntry(self,width=350,textvariable=nome_value,font=("Century Gohtic bold",16),fg_color="transparent")    
        contact_entry= ctk.CTkEntry(self,width=200,textvariable=contact_value,font=("Century Gohtic bold",16),fg_color="transparent")
        age_entry= ctk.CTkEntry(self,width=150,textvariable=age_value,font=("Century Gohtic bold",16),fg_color="transparent")
        address_entry= ctk.CTkEntry(self,width=200,textvariable=address_value,font=("Century Gohtic bold",16),fg_color="transparent")        
    
    #Combobox caixa de genero com  3 opçoes
        gender_combobox = ctk.CTkComboBox(self,value=["Masculino","Feminino","Trasgenero"],font=("Centurty Gothic bold",14),width=150)
        gender_combobox.set("Masculino")
        
    #Entrada de Obcervaçoes
        obs_entry=ctk.CTkTextbox(self,width=450,height=150,font=("Arial",18),border_color="#aaa",border_width=2 ,fg_color="transparent")
        
    #LABEL Caixa de texto titulos
        lb_nome= ctk.CTkLabel(self,text="Nome completo:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_contact= ctk.CTkLabel(self,text="Contato:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_age= ctk.CTkLabel(self,text="Idade:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_gander= ctk.CTkLabel(self,text="Género:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_address= ctk.CTkLabel(self,text="Endereço:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_obs= ctk.CTkLabel(self,text="Observações:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        
     #BOTOES , tamanho cor com interação 
        btn_submit= ctk.CTkButton(self,text="Salvar dados".upper(),command=submit,fg_color="#151",hover_color="#131").place(x=300,y=420)
        btn_submit= ctk.CTkButton(self,text="Limpar campos".upper(),command=clear,fg_color="#555",hover_color="#333").place(x=500,y=420)
        
        
    #Posicionando os elementos na tela
        lb_nome.place(x=50 ,y=120)
        nome_entry.place(x=50 ,y=150)
        
        lb_contact.place(x=450,y=120)
        contact_entry.place(x=450,y=150)
        
        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)
        
        lb_gander.place(x=500,y=190)
        gender_combobox.place(x=500,y=220)
        
        lb_address.place(x=50,y=190)
        address_entry.place(x=50,y=220)
        
        lb_obs.place(x=50,y=260)
        obs_entry.place(x=180 ,y=260)
        
        
    def change_apm(self,nova_aparencia_mode):
        ctk.set_appearance_mode(nova_aparencia_mode)
        
#inicializar a janela
if __name__=="__main__":
    app = App()
    app.mainloop()

#executavel criado com auto-py-to-exe